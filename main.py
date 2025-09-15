import streamlit as st
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import io
import os

# =====================================================================================
#  TRANSLATIONS
# =====================================================================================
translations = {
    'en': {
        'title': "🤖 Automatic Holiday Rescheduler",
        'description': "This tool automates the rescheduling of deliveries in logistics spreadsheets. Just upload your spreadsheet, enter the holiday day, and click 'Reschedule'.",
        'file_uploader_label': "1. Choose your scheduling spreadsheet (.xlsx)",
        'number_input_label': "2. Enter the day of the month that is a holiday (e.g., 20)",
        'button_label': "Reschedule Spreadsheet",
        'spinner_text': "Please wait... Rescheduling tasks...",
        'report_header': "Operation Report:",
        'log_sheet_loaded': "Spreadsheet '{file_name}' loaded successfully.",
        'log_error_read_sheet': "ERROR: Could not read the spreadsheet. Please check if it is the correct file. Details: {error}",
        'log_error_day_not_found': "ERROR: The day {holiday_day} was not found in row 3 of the Delivery columns.",
        'log_holiday_identified': "Holiday identified in the Delivery column: {col_letter}",
        'log_warning_first_day': "Warning: The holiday is the first day of the period. It cannot be anticipated.",
        'log_rescheduled_with_substitution': "Delivery rescheduled (with substitution) from day {holiday_day} to column {col_letter}.",
        'log_rescheduling_complete': "Rescheduling completed. {tasks_moved} tasks were moved.",
        'success_message': "Your spreadsheet has been successfully rescheduled!",
        'download_button_label': "Click here to download the rescheduled spreadsheet",
        'download_file_suffix': "_rescheduled",
        'error_upload_file': "Please upload a spreadsheet before rescheduling."
    },
    'es': {
        'title': "🤖 Reprogramador Automático de Feriados",
        'description': "Esta herramienta automatiza la reprogramación de entregas en planillas de logística. Simplemente suba su planilla, ingrese el día feriado y haga clic en 'Reprogramar'.",
        'file_uploader_label': "1. Elija su planilla de programación (.xlsx)",
        'number_input_label': "2. Ingrese el día del mes que es feriado (ej: 20)",
        'button_label': "Reprogramar Planilla",
        'spinner_text': "Por favor espere... Reprogramando tareas...",
        'report_header': "Reporte de Operación:",
        'log_sheet_loaded': "Planilla '{file_name}' cargada exitosamente.",
        'log_error_read_sheet': "ERROR: No se pudo leer la planilla. Por favor, verifique si es el archivo correcto. Detalles: {error}",
        'log_error_day_not_found': "ERROR: El día {holiday_day} no fue encontrado en la fila 3 de las columnas de Entrega.",
        'log_holiday_identified': "Feriado identificado en la columna de Entrega: {col_letter}",
        'log_warning_first_day': "Advertencia: El feriado es el primer día del período. No se puede anticipar.",
        'log_rescheduled_with_substitution': "Entrega reprogramada (con sustitución) del día {holiday_day} a la columna {col_letter}.",
        'log_rescheduling_complete': "Reprogramación completada. Se movieron {tasks_moved} tareas.",
        'success_message': "¡Su planilla ha sido reprogramada exitosamente!",
        'download_button_label': "Haga clic aquí para descargar la planilla reprogramada",
        'download_file_suffix': "_reprogramada",
        'error_upload_file': "Por favor, suba una planilla antes de reprogramar."
    }
}

# =====================================================================================
#  OUR RESCHEDULING LOGIC (ADAPTED INSIDE A FUNCTION)
# =====================================================================================
def process_spreadsheet(excel_file, holiday_day, texts):
    """
    This function contains our main algorithm.
    It receives the uploaded file, the holiday day, and the language dictionary,
    and returns the modified file and a list of logs.
    """
    logs = []

    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet_name = '01. Calendario SCL Abarrotes'
        sheet = workbook[sheet_name]
        logs.append(texts['log_sheet_loaded'].format(file_name=excel_file.name))
    except Exception as e:
        logs.append(texts['log_error_read_sheet'].format(error=e))
        return None, logs

    delivery_columns = ['AI', 'AJ', 'AK', 'AL', 'AM', 'AN']
    observations_column = 'CT'
    weekday_map = {'L': 1, 'M': 2, 'W': 3, 'J': 4, 'V': 5, 'S': 6, 'D': 7}

    holiday_col_letter = None
    for col_letter in delivery_columns:
        day_in_sheet = sheet[f'{col_letter}3'].value
        if day_in_sheet == holiday_day:
            holiday_col_letter = col_letter
            break
    
    if not holiday_col_letter:
        logs.append(texts['log_error_day_not_found'].format(holiday_day=holiday_day))
        return None, logs
    
    logs.append(texts['log_holiday_identified'].format(col_letter=holiday_col_letter))

    holiday_col_index = column_index_from_string(holiday_col_letter)
    if holiday_col_index == column_index_from_string(delivery_columns[0]):
          logs.append(texts['log_warning_first_day'])
          return None, logs

    previous_col_index = holiday_col_index - 1
    previous_col_letter = get_column_letter(previous_col_index)
    
    tasks_moved = 0
    
    for row_index in range(8, sheet.max_row + 1):
        task_cell = sheet[f'{holiday_col_letter}{row_index}']
        if isinstance(task_cell.value, (int, float)) and 1 <= task_cell.value <= 6:
            destination_cell = sheet[f'{previous_col_letter}{row_index}']
            weekday_initial = sheet[f'{previous_col_letter}6'].value.upper()
            new_weekday_number = weekday_map.get(weekday_initial)
            
            if new_weekday_number:
                destination_cell.value = new_weekday_number
                task_cell.value = None
                log_message = texts['log_rescheduled_with_substitution'].format(
                    holiday_day=holiday_day,
                    col_letter=previous_col_letter
                )
                sheet[f'{observations_column}{row_index}'].value = log_message
                tasks_moved += 1
    
    logs.append(texts['log_rescheduling_complete'].format(tasks_moved=tasks_moved))
    
    return workbook, logs

# =====================================================================================
#  STREAMLIT WEB INTERFACE
# =====================================================================================

language = st.selectbox(
    "Select Language / Seleccione Idioma",
    ('English', 'Español')
)
lang_code = 'en' if language == 'English' else 'es'
texts = translations[lang_code]

st.title(texts['title'])

st.write(texts['description'])

uploaded_file = st.file_uploader(
    texts['file_uploader_label'],
    type=['xlsx']
)

holiday_day = st.number_input(
    texts['number_input_label'],
    min_value=1, 
    max_value=31, 
    step=1,
    value=20
)

if st.button(texts['button_label']):
    if uploaded_file is not None:
        with st.spinner(texts['spinner_text']):
            modified_workbook, logs = process_spreadsheet(uploaded_file, int(holiday_day), texts)

        st.subheader(texts['report_header'])
        for log in logs:
            st.info(log)
        
        if modified_workbook:
            output = io.BytesIO()
            modified_workbook.save(output)
            output.seek(0)
            
            st.success(texts['success_message'])
            
            original_filename = uploaded_file.name
            base_name, extension = os.path.splitext(original_filename)
            new_filename = f"{base_name}{texts['download_file_suffix']}{extension}"
            
            st.download_button(
                label=texts['download_button_label'],
                data=output,
                file_name=new_filename,
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
    else:
        st.error(texts['error_upload_file'])